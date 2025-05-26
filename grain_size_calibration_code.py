import math
import numpy as np
import cv2
from matplotlib import pyplot as plt
import pandas as pd
from scipy import ndimage
from skimage import io, color, measure
import os
import seaborn as sns
from scipy.stats import gaussian_kde
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from finddisc import crop_disk_from_image
# from #jeszcze nie ma szukania ziaren

class Grains:
    def __init__(self):
        self.file_name_L, self.index_L, self.cluster_orientation_L, self.cluster_equivalent_diameter_area_L, self.cluster_perimeter_L, self.x_L, self.y_L= [], [], [], [], [], [], []

        folders = ['output/detected_disc', 'output/detected_grains', 'output/for_user', 'input']
        for folder in folders:
            os.makedirs(folder, exist_ok=True)

    def process_image(self ,file_name, min_sieves, max_sieves, i):

        img = cv2.imread('input/' + file_name)
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        _, binary = cv2.threshold(gray, 10, 200, cv2.THRESH_BINARY_INV+cv2.THRESH_OTSU)
        contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        largest_contour = max(contours, key=cv2.contourArea)
        (x, y), radius = cv2.minEnclosingCircle(largest_contour)
        area = cv2.contourArea(largest_contour)
        radius_from_area = (area/math.pi)**.5
        radius_for_calculations = min(radius, radius_from_area)

        um_per_pixel = (10_000)/(2*radius_for_calculations)
        um2_per_pixel = (math.pi*5_000**2)/area

        cv2.circle(img, (int(x), int(y)), int(radius), (0, 255, 0), 2)
        cv2.drawContours(img, [largest_contour], -1,(0, 0, 255), 2)
        cv2.imwrite(f'output/detected_disc/detected_disc_{i}.jpg', img)

        img = cv2.imread('input/' + file_name, 0)
        ret, thresh = cv2.threshold(img, 150, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        cv2.imwrite(os.path.join('output', 'detected_grains', f'thresh_obraz_{i}.jpg'), thresh)
        kernel = np.ones((3, 3), np.uint8)
        eroded = cv2.erode(thresh, kernel, iterations=1)
        cv2.imwrite(os.path.join('output', 'detected_grains', f'eroded_obraz_{i}.jpg'), eroded)
        dilated = cv2.dilate(eroded, kernel, iterations=1)
        cv2.imwrite(os.path.join('output', 'detected_grains', f'dilated_obraz_{i}.jpg'), dilated)
        mask = dilated == 255
        s = np.ones((3, 3), dtype=int)
        label_mask, num_labels = ndimage.label(mask, structure=s)

        img_color = color.label2rgb(label_mask, bg_label=0)
        img_color = (img_color * 255).astype(np.uint8)
        cv2.imwrite(os.path.join('output', 'detected_grains', f'wykryte_ziarna_{i}.jpg'),
                    cv2.cvtColor(img_color, cv2.COLOR_RGB2BGR))

        clusters = measure.regionprops(label_mask, img)

        rejected_list = ['']*len(clusters)
        Recection_dictionary = []

        Recection_dictionary.append(['Q','Equivalent diameter can be up to 150% of larger sieve size see Fig 6 in Poreba et al. (2022), removes overlying'])
        for index, cluster in enumerate(clusters):
            area_um2 = cluster.area * um2_per_pixel
            if 2*(area_um2/math.pi)**.5 >= (1.5 * max_sieves):
                rejected_list[index] += 'Q'

        Recection_dictionary.append(['W','Equivalent diameter should above 50% of smaller sieve size see Fig 6 in Poreba et al. (2022), removes broken grains'])
        for index, cluster in enumerate(clusters):
            area_um2 = cluster.area * um2_per_pixel
            if 2*(area_um2/math.pi)**.5 <= (0.5 * min_sieves):
                rejected_list[index] += 'W'

        Recection_dictionary.append(['E','Roundness needs to be larger than 0.3 see Fig 6 in Poreba et al. (2022), removes overlying'])
        for index, cluster in enumerate(clusters):
            if 4*math.pi*cluster.area/(cluster.perimeter**2) < .6:
                rejected_list[index] += f'E{4*math.pi*cluster.area/(cluster.perimeter**2):.1f}'

        img_color = cv2.cvtColor(img_color, cv2.COLOR_RGB2BGR)
        img = cv2.imread('input/' + file_name)
        cv2.circle(img, (int(x), int(y)), int(radius), (0, 0, 255), 2)
        cv2.drawContours(img, [largest_contour], -1, (0, 255, 0), 2)

        for index, region_label in enumerate(range(1, num_labels + 1)):
            single_mask = (label_mask == region_label).astype(np.uint8) * 255
            contours, _ = cv2.findContours(single_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            if rejected_list[index]:
                cv2.drawContours(img, contours, -1, (0, 0, 255), 1)
            else:
                cv2.drawContours(img, contours, -1, (0, 255, 0), 1)

        for index, cluster in enumerate(clusters):
            y, x = map(int, cluster.centroid)
            text_to_display = str(cluster.label) + rejected_list[index]
            if rejected_list[index]:
                cv2.putText(img, text_to_display, (x, y), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 1)
            else:
                cv2.putText(img, text_to_display, (x, y), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 1)
                self.file_name_L.append(file_name)
                self.index_L.append(index)
                self.cluster_orientation_L.append(cluster.orientation)
                self.cluster_equivalent_diameter_area_L.append(cluster.equivalent_diameter_area*um_per_pixel)
                self.cluster_perimeter_L.append(cluster.perimeter*um_per_pixel)
                self.x_L.append(x)
                self.y_L.append(y)

        cv2.imwrite(f'output/for_user/{file_name[:-4]}_processed.jpg', img)
        return

    def statistics_and_save(self):
        df = pd.DataFrame({
            'file_name': self.file_name_L,
            'index': self.index_L,
            'x (pixels)': self.x_L,
            'y (pixels)': self.y_L,
            'diameter (um)': self.cluster_equivalent_diameter_area_L,
            'perimeter (um)': self.cluster_perimeter_L,
        })

        wb = Workbook()
        ws = wb.active
        ws.title = "Grain Data"

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        last_row = len(df) + 1
        ws["H1"].value = f"AVERAGE (um)"
        ws["H2"].value = f"MEDIAN (um)"
        ws["H3"].value = f"STDEV (um)"

        ws["I1"].value = f"=AVERAGE(E2:E{last_row})"
        ws["I2"].value = f"=MEDIAN(E2:E{last_row})"
        ws["I3"].value = f"=STDEV(E2:E{last_row})"

        wb.save("output/for_user/grain_data.xlsx")

        plt.figure(figsize=(10, 6))
        sns.histplot(data=df, x='diameter (um)', hue='file_name', bins=100, multiple='stack', edgecolor='black', palette='tab10', stat='count')

        data = np.array(df['diameter (um)'])
        median = np.median(data)
        lower_68, upper_68 = np.percentile(data, [16, 84])
        lower_95, upper_95 = np.percentile(data, [2.5, 97.5])

        kde = gaussian_kde(data)
        x_vals = np.linspace(data.min(), data.max(), 500)
        kde_vals = kde(x_vals)
        scaled_kde = kde_vals * len(data) * (x_vals[1] - x_vals[0])
        plt.plot(x_vals, scaled_kde, color='black', linestyle='-', label='Density')

        plt.axvline(median, color='black', linestyle='--', linewidth=2, label='Median')
        plt.axvspan(lower_68, upper_68, color='gray', alpha=0.3, label='68.2% interval')
        plt.axvspan(lower_95, upper_95, color='gray', alpha=0.2, label='95.4% interval')

        plt.title("Stacked Histogram of Equivalent Diameters")
        plt.xlabel("Equivalent Diameter (area-based), um")
        plt.ylabel("Count")
        plt.legend()
        plt.tight_layout()
        plt.savefig("output/for_user/stacked_histogram_diameter.png", dpi=300)
        plt.show()

if __name__ == '__main__':

    min_sieves, max_sieves = 150, 250

    list_jpg = sorted([f for f in os.listdir('input/') if f.lower().endswith('.jpg')])
    Grains=Grains()
    for i, single_image in enumerate(list_jpg, 1):
        print(f'Processing - {single_image}')
        #find disc
        crop_disk_from_image(single_image)
        Grains.process_image(single_image, min_sieves, max_sieves, i)
    Grains.statistics_and_save()
    print('the end')